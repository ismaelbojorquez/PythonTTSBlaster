import asyncio
import csv
import io
import sqlite3
from datetime import date
from pathlib import Path

import pytest
from conftest import campaign
from openpyxl import load_workbook

from blaster.analytics import Analytics, Filters
from blaster.reports import cdr_csv, excel_report
from blaster.store import Store


async def until(predicate, timeout=3):
    async with asyncio.timeout(timeout):
        while not predicate():
            await asyncio.sleep(0.002)


async def test_call_trace_records_two_legs_transfer_and_hangup_actor(engine):
    cid = campaign(engine)
    jid = engine.store.jobs(cid)[0]["id"]
    engine.start_campaign(cid)
    await until(lambda: jid in engine.sessions and engine.sessions[jid].state == "menu")
    engine.simulate(jid, "2")
    await until(lambda: engine.sessions[jid].state == "bridged")
    engine.simulate(jid, "agent_hangup")
    await until(lambda: not engine.sessions)

    analytics = Analytics(Path(engine.store.db.execute("PRAGMA database_list").fetchone()[2]))
    row = analytics.calls(Filters(mode="simulation"))["items"][0]
    assert row["coverage"] == "measured"
    assert row["customer_answered_at"] and row["agent_answered_at"]
    assert row["transfer_actor"] == "customer"
    assert row["bridge_seconds"] >= 0
    assert row["end_actor"] == "agent"
    assert row["customer_trunk_id"] == "default"
    assert row["customer_trunk_name"] == "Troncal principal"
    assert row["agent_trunk_id"] == "default"
    assert row["agent_trunk_name"] == "Troncal principal"
    detail = analytics.detail(jid)
    assert detail["attempts"][0]["customer_trunk_id"] == "default"
    assert detail["attempts"][0]["customer_trunk_name"] == "Troncal principal"
    assert {leg["trunk_name"] for leg in detail["legs"]} == {"Troncal principal"}
    summary = analytics.summary(Filters(mode="simulation"))
    assert summary["counts"]["attempted"] == summary["counts"]["answered"] == 1
    assert summary["counts"]["bridged"] == 1
    assert summary["answer_rate"] == summary["transfer_rate"] == 1

    rows, summary, events = analytics.report_data(Filters(mode="simulation"), 100)
    rows[0]["contact_name"] = "=HYPERLINK(\"https://example.invalid\",\"click\")"
    csv_payload = cdr_csv(rows)
    assert csv_payload.startswith(b"\xef\xbb\xbf")
    assert "'=HYPERLINK" in csv_payload.decode("utf-8-sig")
    exported = next(csv.DictReader(io.StringIO(csv_payload.decode("utf-8-sig"))))
    assert exported["Troncal de salida"] == "Troncal principal"
    assert exported["ID troncal de salida"] == "default"
    assert exported["Troncal de transferencia"] == "Troncal principal"
    assert exported["ID troncal de transferencia"] == "default"
    payload = excel_report(rows, summary, events, Filters(mode="simulation"))
    assert payload[:2] == b"PK"
    workbook = load_workbook(io.BytesIO(payload))
    assert len(workbook.sheetnames) == 8
    headers = {cell.value: cell.column for cell in workbook["CDRs"][1]}
    assert workbook["CDRs"].cell(2, headers["Nombre en contacto"]).data_type == "s"
    assert workbook["CDRs"].cell(2, headers["Nombre en contacto"]).value == rows[0]["contact_name"]
    assert workbook["CDRs"].cell(2, headers["Teléfono cliente"]).data_type == "s"
    assert workbook["CDRs"].cell(2, headers["Troncal de salida"]).value == "Troncal principal"
    assert workbook["CDRs"].cell(2, headers["ID troncal de salida"]).value == "default"
    assert workbook["CDRs"].cell(2, headers["Cliente conectado (s)"]).data_type == "n"
    leg_headers = {cell.value: cell.column for cell in workbook["Tramos"][1]}
    assert workbook["Tramos"].cell(2, leg_headers["Nombre troncal"]).value == "Troncal principal"
    assert workbook["Tramos"].cell(2, leg_headers["trunk_id"]).value == "default"
    assert len(workbook["Resumen"]._charts) == 2

    english_csv = next(
        csv.DictReader(io.StringIO(cdr_csv(rows, "en").decode("utf-8-sig")))
    )
    assert english_csv["Outbound provider"] == "Troncal principal"
    assert english_csv["Operation type"] == "Test"
    english_book = load_workbook(
        io.BytesIO(excel_report(rows, summary, events, Filters(mode="simulation"), "en")),
        read_only=True,
    )
    assert "Overview" in english_book.sheetnames
    assert "Call legs" in english_book.sheetnames
    assert "Account" in [cell.value for cell in next(english_book["CDRs"].iter_rows())]
    english_book.close()
    with pytest.raises(ValueError, match="no se exportaron datos parciales"):
        analytics.report_data(Filters(mode="simulation"), 0)


def test_migration_preserves_legacy_data_and_does_not_invent_metrics(tmp_path):
    path = tmp_path / "legacy.sqlite3"
    db = sqlite3.connect(path)
    db.executescript("""
        CREATE TABLE campaigns (id TEXT PRIMARY KEY,name TEXT,template TEXT,
          agent_number TEXT,status TEXT,created_at TEXT,mode TEXT);
        CREATE TABLE jobs (id TEXT PRIMARY KEY,campaign_id TEXT,phone TEXT,variables TEXT,
          message TEXT,status TEXT,detail TEXT,started_at TEXT,ended_at TEXT,updated_at TEXT);
        CREATE TABLE events (id INTEGER PRIMARY KEY,job_id TEXT,status TEXT,detail TEXT,
          created_at TEXT);
        INSERT INTO campaigns VALUES('c','Histórica','Hola','52550000','completed',
          '2025-01-01T00:00:00+00:00','sip');
        INSERT INTO jobs VALUES('j','c','52551111','{}','Hola','completed','Terminó',
          '2025-01-01T00:00:00+00:00','2025-01-01T00:01:00+00:00',
          '2025-01-01T00:01:00+00:00');
    """)
    db.close()
    store = Store(path)
    store.close()
    Store(path).close()  # an already migrated database is not backed up again
    result = Analytics(path).summary(Filters(date_from=date(2024, 12, 31), mode="sip"))
    assert result["counts"]["total"] == result["counts"]["legacy"] == 1
    assert result["counts"]["answered"] == result["counts"]["attempted"] == 0
    assert result["answer_rate"] is None
    assert len(list(tmp_path.glob("*.before-analytics-*.bak"))) == 1
    empty = Analytics(path).summary(Filters(date_from=date(2025, 1, 1), mode="sip"))
    assert empty["counts"]["total"] == 0  # UTC midnight belongs to the prior local date
