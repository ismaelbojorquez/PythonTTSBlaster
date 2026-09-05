import asyncio
import io
import sqlite3

from conftest import remove_retries_schema
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from test_engine import until

from blaster.agent_pool import AgentPool
from blaster.analytics import Analytics, Filters
from blaster.config import Settings
from blaster.models import CampaignInput, Contact
from blaster.reports import cdr_csv, excel_report
from blaster.store import Store
from blaster.web import create_app

A, B, C = "525512345671", "525512345672", "525512345673"


def make_campaign(engine, numbers=(A, B), count=2, wait=0.5, strategy="round_robin"):
    return engine.store.create_campaign(
        CampaignInput(
            name="Pool de prueba",
            template="Hola",
            agent_numbers=list(numbers),
            agent_strategy=strategy,
            agent_pool_wait=wait,
            contacts=[
                Contact(phone=f"5255550000{i:02}", credit_id=f"CRED-{i:03}")
                for i in range(count)
            ],
        )
    )


async def request_transfers(engine, cid):
    jobs = engine.store.jobs(cid)
    engine.start_campaign(cid)
    await until(
        lambda: all(
            j["id"] in engine.sessions and engine.sessions[j["id"]].state == "menu" for j in jobs
        )
    )
    for job in jobs:
        engine.simulate(job["id"], "2")
    return [j["id"] for j in jobs]


async def test_simultaneous_transfers_reserve_distinct_numbers_until_hangup(engine):
    engine.settings.agent_timeout = 0.5
    cid = make_campaign(engine)
    ids = await request_transfers(engine, cid)
    await until(lambda: all(engine.sessions[j].state == "bridged" for j in ids))
    actual = {j: engine.sessions[j].agent.number for j in ids}
    assert set(actual.values()) == {A, B}
    assert len(engine.agent_pool.busy) == 2
    assert engine.snapshot()["channels_in_use"] == 4
    for jid in ids:
        engine.simulate(jid, "agent_hangup")
    await until(lambda: not engine.sessions)
    assert not engine.agent_pool.busy
    analytics = Analytics(engine.settings.data_dir / "test.sqlite3")
    rows, summary, events = analytics.report_data(Filters(mode="simulation"), 100)
    assert {r["id"]: r["agent_number"] for r in rows} == actual
    assert all(r["agent_strategy"] == "round_robin" for r in rows)
    assert all(r["agent_pool_wait_seconds"] is not None for r in rows)
    assert B in cdr_csv(rows).decode("utf-8-sig")
    workbook = load_workbook(
        io.BytesIO(excel_report(rows, summary, events, Filters(mode="simulation")))
    )
    headers = {cell.value: cell.column for cell in workbook["CDRs"][1]}
    assert {
        workbook["CDRs"].cell(i, headers["Número agente"]).value for i in (2, 3)
    } == {A, B}
    assert {
        e["data"]["number"]
        for e in analytics.detail(ids[0])["events"]
        if e["kind"] == "agent_selected"
    } == {actual[ids[0]]}


async def test_single_number_waits_then_reuses_only_after_first_call_closes(engine):
    cid = make_campaign(engine, numbers=(A,), wait=2)
    ids = await request_transfers(engine, cid)
    await until(lambda: {engine.sessions[j].state for j in ids} == {"bridged", "agent_waiting"})
    first = next(j for j in ids if engine.sessions[j].state == "bridged")
    second = next(j for j in ids if j != first)
    assert engine.sessions[second].agent is None
    assert sum(leg.role == "agent" for leg in engine.telephony.legs.values()) == 1
    engine.simulate(first, "hangup")
    await until(lambda: engine.sessions[second].state == "bridged")
    assert engine.sessions[second].agent.number == A
    assert engine.agent_pool.busy == {A: second}
    engine.simulate(second, "agent_hangup")
    await until(lambda: not engine.sessions)
    assert not engine.agent_pool.busy and not engine.agent_pool.pending


async def test_full_pool_pauses_new_originations_and_resumes_on_release(engine):
    cid = make_campaign(engine, numbers=(A,), count=2)
    engine.agent_pool.busy[A] = "existing-transfer"
    engine.start_campaign(cid)

    await until(lambda: engine.snapshot()["origination_paused"])
    await asyncio.sleep(0.3)
    assert not engine.sessions
    assert {job["status"] for job in engine.store.jobs(cid)} == {"queued"}
    pool = engine.snapshot()["agent_pool"]
    assert (pool["total"], pool["busy"], pool["free"]) == (1, 1, 0)
    paused = engine.store.db.execute(
        "SELECT COUNT(*) FROM audit WHERE action='campaign.capacity_paused' AND target=?",
        (cid,),
    ).fetchone()[0]
    assert paused == 1

    engine.agent_pool.release("existing-transfer")
    await until(lambda: bool(engine.sessions))
    assert not engine.snapshot()["origination_paused"]
    resumed = engine.store.db.execute(
        "SELECT COUNT(*) FROM audit WHERE action='campaign.capacity_resumed' AND target=?",
        (cid,),
    ).fetchone()[0]
    assert resumed == 1
    await engine.stop_campaign(cid)


async def test_transferred_call_holds_next_contact_until_agent_is_free(engine):
    cid = make_campaign(engine, numbers=(A,), count=3, wait=2)
    jobs = engine.store.jobs(cid)
    first, second, third = [job["id"] for job in jobs]
    engine.start_campaign(cid)
    await until(
        lambda: first in engine.sessions
        and second in engine.sessions
        and engine.sessions[first].state == "menu"
        and engine.sessions[second].state == "menu"
    )

    engine.simulate(first, "2")
    await until(lambda: engine.sessions[first].state == "bridged")
    engine.simulate(second, "hangup")
    await until(lambda: second not in engine.sessions)
    await until(lambda: engine.snapshot()["origination_paused"])
    assert third not in engine.sessions
    third_job = next(job for job in engine.store.jobs(cid) if job["id"] == third)
    assert third_job["status"] == "queued"

    engine.simulate(first, "agent_hangup")
    await until(lambda: third in engine.sessions)
    assert not engine.snapshot()["origination_paused"]
    await engine.stop_campaign(cid)


async def test_legacy_national_agent_matches_its_international_reservation(engine):
    engine.settings.sip.dial_format = "mexico_52"
    cid = make_campaign(engine, numbers=("5512345671",), count=1)
    ids = await request_transfers(engine, cid)
    await until(lambda: engine.sessions[ids[0]].state == "bridged")
    reservation = engine.snapshot()["agent_pool"]["reservations"][0]
    assert reservation["number"] == A
    assert reservation["configured_number"] == "5512345671"
    engine.simulate(ids[0], "hangup")
    await until(lambda: not engine.sessions)
    assert not engine.agent_pool.aliases


async def test_full_pool_timeout_records_wait_without_dialing_an_occupied_number(engine):
    cid = make_campaign(engine, numbers=(A,), count=1, wait=0.02)
    jobs = engine.store.jobs(cid)
    engine.start_campaign(cid)
    await until(
        lambda: jobs[0]["id"] in engine.sessions
        and engine.sessions[jobs[0]["id"]].state == "menu"
    )
    engine.agent_pool.busy[A] = "other-session"
    engine.simulate(jobs[0]["id"], "2")
    ids = [jobs[0]["id"]]
    await until(lambda: not engine.sessions)
    assert engine.store.jobs(cid)[0]["status"] == "failed"
    assert "pool" in engine.store.jobs(cid)[0]["detail"]
    row = Analytics(engine.settings.data_dir / "test.sqlite3").detail(ids[0])
    assert row["agent_number"] is None and row["agent_pool_wait_seconds"] >= 0.02
    assert not any(leg["role"] == "agent" for leg in row["legs"])
    assert not engine.agent_pool.pending


async def test_hanging_up_while_queued_cancels_the_reservation_request(engine):
    cid = make_campaign(engine, numbers=(A,), count=1, wait=2)
    jobs = engine.store.jobs(cid)
    engine.start_campaign(cid)
    jid = jobs[0]["id"]
    await until(lambda: jid in engine.sessions and engine.sessions[jid].state == "menu")
    engine.agent_pool.busy[A] = "other-session"
    engine.simulate(jid, "2")
    await until(lambda: engine.sessions[jid].state == "agent_waiting")
    engine.simulate(jid, "hangup")
    await until(lambda: not engine.sessions)
    assert not engine.agent_pool.pending
    assert engine.agent_pool.busy == {A: "other-session"}


async def test_stop_during_cps_wait_releases_claims_without_creating_agent_legs(engine):
    cid = make_campaign(engine)
    engine.start_campaign(cid)
    jobs = engine.store.jobs(cid)
    await until(
        lambda: all(
            j["id"] in engine.sessions and engine.sessions[j["id"]].state == "menu" for j in jobs
        )
    )
    engine.next_dial = asyncio.get_running_loop().time() + 10
    for job in jobs:
        engine.simulate(job["id"], "2")
    await until(lambda: len(engine.agent_pool.busy) == 2)
    assert all(s.agent is None for s in engine.sessions.values())
    await engine.stop_campaign(cid)
    assert not engine.agent_pool.busy and not engine.agent_pool.pending


async def test_selection_policies_and_round_robin_cursor_survive_recreating_pool(
    engine, monkeypatch
):
    cid = make_campaign(engine, numbers=(A, B, C), count=1)
    pool = engine.agent_pool

    async def take(owner, strategy="round_robin", timeout=0):
        return await pool.acquire(owner, cid, [A, B, C], strategy, timeout, lambda: None)

    assert await take("one") == A
    pool.release("one")
    pool = AgentPool(engine.store)
    assert await take("two") == B
    assert await take("three") == C
    # B and C are held; priority uses the only free destination A.
    assert await take("four", "priority") == A
    assert await take("full", "random") is None
    pool.release("two")
    monkeypatch.setattr("blaster.agent_pool.random.choice", lambda free: free[-1])
    assert await take("random", "random") == B
    await pool.close()


async def test_capacity_uses_the_configured_number_behind_a_dial_alias(engine):
    pool = engine.agent_pool
    pool.busy[A] = "held"
    pool.aliases[A] = "5512345671"
    capacity = pool.availability(["5512345671"])
    assert capacity == {"total": 1, "busy": 1, "free": []}


async def test_unconfirmed_hangup_does_not_release_destination(engine):
    cid = make_campaign(engine, numbers=(A,), count=1)
    pool = engine.agent_pool
    assert await pool.acquire("held", cid, [A], "priority", 0, lambda: None) == A

    class Leg:
        closed = asyncio.Event()

    leg = Leg()
    pool.release_after_close("held", leg)
    assert await pool.acquire("next", cid, [A], "priority", 0, lambda: None) is None
    leg.closed.set()
    await until(lambda: not pool.busy)
    assert await pool.acquire("next", cid, [A], "priority", 0, lambda: None) == A


async def test_late_sip_disconnection_keeps_leg_reachable_until_pool_release(engine, monkeypatch):
    from blaster.telephony.pjsua import PJSUALeg, PJSUATelephony

    phone = PJSUATelephony(Settings())

    async def command(*args):
        return None

    async def expired(awaitable, timeout):
        awaitable.close()
        raise TimeoutError

    phone.command = command
    leg = PJSUALeg(A, "agent", phone)
    phone.legs[leg.id] = leg
    engine.agent_pool.busy[A] = "held"
    with monkeypatch.context() as patch:
        patch.setattr("blaster.telephony.pjsua.asyncio.wait_for", expired)
        await leg.hangup()
    engine.agent_pool.release_after_close("held", leg)
    assert not leg.closed.is_set() and leg.id in phone.legs
    assert engine.agent_pool.busy[A] == "held"
    phone._event(leg.id, "closed", (200, "OK"))
    await until(lambda: not engine.agent_pool.busy)
    assert leg.id not in phone.legs


def test_campaign_pool_and_template_validation_and_persistence(tmp_path):
    payload = {
        "name": "Pool",
        "country": "MX",
        "agent_numbers_text": "5512345671\n+525512345672",
        "agent_strategy": "random",
        "agent_pool_wait": 12,
        "template": "Hola",
        "csv_text": "Credito,telefono\nPOOL-1,5512345600",
    }
    with TestClient(create_app(Settings(data_dir=tmp_path, auth={"enabled": False}))) as client:
        preview = client.post("/api/preview", json=payload)
        assert preview.status_code == 200, preview.text
        assert preview.json()["agent_numbers"] == [A, B]
        response = client.post("/api/campaigns", json=payload)
        assert response.status_code == 201, response.text
        cid = response.json()["id"]
        campaign = client.get("/api/campaigns").json()[0]
        assert campaign["agent_numbers"] == [A, B] and campaign["agent_strategy"] == "random"
        assert campaign["agent_pool_wait"] == 12
        for change in [
            {"agent_numbers_text": ""},
            {"agent_numbers_text": "5512345671\n525512345671"},
            {"agent_strategy": "invalid"},
            {"agent_pool_wait": 301},
        ]:
            assert client.post("/api/campaigns", json={**payload, **change}).status_code == 422
        template = {
            "name": "Pool",
            "message": "Hola",
            "agent_country": "MX",
            "agent_numbers_text": payload["agent_numbers_text"],
            "agent_strategy": "priority",
            "agent_pool_wait": 15,
        }
        assert client.post("/api/manage/templates", json=template).status_code == 200
        saved = client.get("/api/manage/templates").json()[0]
        assert saved["agent_numbers"] == [A, B] and saved["agent_pool_wait"] == 15
        assert saved["agent_strategy"] == "priority"
        assert len(saved["agent_numbers_national"]) == 2
    store = Store(tmp_path / "blaster.sqlite3")
    assert store.campaign(cid)["agent_numbers"] == [A, B]
    store.close()


def test_v3_migration_keeps_existing_campaign_as_single_number_pool(tmp_path):
    path = tmp_path / "v3.sqlite3"
    Store(path).close()
    with sqlite3.connect(path) as db:
        remove_retries_schema(db)
        db.execute("DROP TABLE campaign_copies")
        for table in ("campaigns", "templates"):
            for column in ("agent_numbers", "agent_strategy", "agent_pool_wait"):
                db.execute(f"ALTER TABLE {table} DROP COLUMN {column}")
        db.execute("ALTER TABLE campaigns DROP COLUMN agent_cursor")
        for column in ("agent_selected_number", "agent_strategy", "agent_pool_wait_seconds"):
            db.execute(f"ALTER TABLE call_records DROP COLUMN {column}")
        db.execute("PRAGMA user_version=3")
        db.execute(
            "INSERT INTO campaigns VALUES"
            "('old','Anterior','Hola',?,'draft','2026','sip','MX','MX')",
            (A,),
        )
    for _ in range(2):
        store = Store(path)
        campaign = store.campaign("old")
        assert campaign["agent_numbers"] == [A] and campaign["agent_strategy"] == "round_robin"
        assert campaign["agent_cursor"] == 0
        store.close()
    assert len(list(tmp_path.glob("*.before-agent-pool-*.bak"))) == 1
