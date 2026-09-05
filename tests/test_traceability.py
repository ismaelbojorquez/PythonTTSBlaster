import io
import sqlite3
from zipfile import ZipFile

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from blaster.config import Settings
from blaster.models import CampaignInput, Contact
from blaster.store import Store
from blaster.web import create_app


def payload(name, credit, phone):
    return {
        "name": name,
        "template": "Aviso del credito {credito}",
        "agent_number": "5550000999",
        "csv_text": f"Credito,Telefono,nombre\n{credit},{phone},Ana",
        "country": "MX",
    }


def finish_all(app):
    with sqlite3.connect(app.state.analytics.path) as db:
        db.execute(
            "UPDATE jobs SET status='completed',detail='Prueba terminada',"
            "started_at='2026-09-04T12:00:00+00:00',"
            "ended_at='2026-09-04T12:00:05+00:00'"
        )
        db.execute(
            "INSERT INTO call_legs(id,job_id,role,number,created_at,invite_at,ended_at,trunk_id) "
            "SELECT lower(hex(randomblob(16))),id,'customer',phone,started_at,started_at,"
            "ended_at,'default' FROM jobs WHERE NOT EXISTS "
            "(SELECT 1 FROM call_legs WHERE call_legs.job_id=jobs.id AND role='customer')"
        )


def test_credit_is_required_and_survives_retry_and_copy(tmp_path):
    store = Store(tmp_path / "trace.sqlite3")
    campaign = CampaignInput(
        name="Trazable",
        template="Credito {credito}",
        agent_number="999",
        contacts=[Contact(phone="123", credit_id="CRED-001")],
        retry_policy={"max_attempts": 2, "delay_seconds": 1, "outcomes": ["no_answer"]},
    )
    cid = store.create_campaign(campaign)
    store.set_campaign_status(cid, "running")
    job = store.jobs(cid)[0]
    assert job["credit_id"] == "CRED-001" and job["message"] == "Credito CRED-001"
    with store.db:
        store.db.execute(
                "INSERT INTO call_records(job_id,version,started_at,finalized_at,"
                "amd_enabled,amd_verdict) "
                "VALUES(?,1,'2026-09-04T00:00:00+00:00',"
                "'2026-09-04T00:00:01+00:00',0,'unmeasured')",
            (job["id"],),
        )
    store.transition(job["id"], "no_answer")
    retry = store.retries.plan(job["id"])
    saved_credit = store.db.execute(
        "SELECT credit_id FROM jobs WHERE id=?", (retry,)
    ).fetchone()[0]
    assert saved_credit == "CRED-001"
    actor = {"id": "admin", "username": "admin"}
    fingerprint = store.history.fingerprint(cid, "duplicate", "Copia", "", actor)
    copy_id = store.history.copy(
        cid, "duplicate", "Copia", "", "00000000-0000-4000-8000-000000000001",
        actor, "simulation", fingerprint,
    )
    assert store.jobs(copy_id)[0]["credit_id"] == "CRED-001"
    with pytest.raises(ValueError, match="Credito"):
        Contact(phone="123", credit_id="   ")
    store.close()


async def test_campaign_cannot_start_with_a_legacy_missing_identifier(engine):
    cid = engine.store.create_campaign(
        CampaignInput(
            name="Incompleta", template="Hola", agent_number="999",
            contacts=[Contact(phone="123", credit_id="TEMP")],
        )
    )
    with engine.store.db:
        engine.store.db.execute("UPDATE jobs SET credit_id='' WHERE campaign_id=?", (cid,))
    with pytest.raises(ValueError, match="sin Credito o Telefono"):
        engine.start_campaign(cid)
    assert engine.store.campaign(cid)["status"] == "draft"


def test_v6_migration_keeps_missing_historical_credit_explicit(tmp_path):
    path = tmp_path / "v6.sqlite3"
    store = Store(path)
    cid = store.create_campaign(
        CampaignInput(
            name="Anterior", template="Hola", agent_number="999",
            contacts=[Contact(phone="123", credit_id="TEMP")],
        )
    )
    with store.db:
        store.db.execute("DROP INDEX jobs_credit")
        store.db.execute("DROP INDEX jobs_phone")
        store.db.execute("ALTER TABLE jobs DROP COLUMN credit_id")
        store.db.execute("PRAGMA user_version=6")
    store.close()
    reopened = Store(path)
    assert reopened.db.execute("PRAGMA user_version").fetchone()[0] == 8
    assert reopened.jobs(cid)[0]["credit_id"] == ""
    assert reopened.missing_identifiers(cid) == 1
    actor = {"id": "admin", "username": "admin"}
    fingerprint = reopened.history.fingerprint(cid, "duplicate", "Copia", "", actor)
    with pytest.raises(ValueError, match="sin Credito o Telefono"):
        reopened.history.copy(
            cid, "duplicate", "Copia", "", "00000000-0000-4000-8000-000000000002",
            actor, "simulation", fingerprint,
        )
    reopened.close()
    assert len(list(tmp_path.glob("*.before-traceability-*.bak"))) == 1


def test_exact_trace_report_and_recording_bundle(tmp_path):
    app = create_app(
        Settings(data_dir=tmp_path, mode="simulation", auth={"enabled": False})
    )
    with TestClient(app) as client:
        for args in (
            ("Primera", "CRED-001", "5512345678"),
            ("Segunda", "CRED-001", "5587654321"),
            ("Otro crédito", "CRED-002", "5512345678"),
        ):
            response = client.post("/api/campaigns", json=payload(*args))
            assert response.status_code == 201, response.text
        finish_all(app)
        campaigns = client.get("/api/campaigns").json()
        first = next(
            row for campaign in campaigns
            for row in client.get(f"/api/campaigns/{campaign['id']}/jobs").json()
            if row["credit_id"] == "CRED-001"
        )
        recording_dir = app.state.engine.recordings.directory
        recording_dir.mkdir(parents=True, exist_ok=True)
        audio = recording_dir / f"{first['id']}.ogg"
        audio.write_bytes(b"OggS-test-audio")
        with sqlite3.connect(app.state.analytics.path) as db:
            db.execute(
                "INSERT INTO recordings(job_id,status,filename,started_at,ended_at,evidence,"
                "size_bytes,duration_seconds) VALUES(?,'ready',?,?,?,'test',?,1.5)",
                (
                    first["id"], f"{first['id']}.ogg", first["started_at"], first["ended_at"],
                    audio.stat().st_size,
                ),
            )

        result = client.get("/api/traceability", params={"by": "credit", "query": "CRED-001"})
        assert result.status_code == 200, result.text
        data = result.json()
        assert data["total"] == 2 and data["metrics"]["campaigns"] == 2
        assert data["metrics"]["recordings"] == 1
        assert {row["credit_id"] for row in data["items"]} == {"CRED-001"}
        assert {row["customer_trunk_id"] for row in data["items"]} == {"default"}
        assert {row["customer_trunk_name"] for row in data["items"]} == {
            "Troncal principal"
        }

        by_phone = client.get(
            "/api/traceability", params={"by": "phone", "query": "+52 (55) 1234-5678"}
        ).json()
        assert by_phone["total"] == 2
        assert {row["credit_id"] for row in by_phone["items"]} == {"CRED-001", "CRED-002"}

        by_national_phone = client.get(
            "/api/traceability",
            params={"by": "phone", "query": "55 1234-5678", "country": "MX"},
        ).json()
        assert by_national_phone["total"] == 2
        assert {row["phone"] for row in by_national_phone["items"]} == {"525512345678"}

        phone_report = client.get(
            "/api/traceability/report.xlsx",
            params={"by": "phone", "query": "5512345678", "country": "MX"},
        )
        assert phone_report.status_code == 200, phone_report.text

        phone_bundle = client.get(
            "/api/traceability/bundle.zip",
            params={"by": "phone", "query": "5512345678", "country": "MX"},
        )
        assert phone_bundle.status_code == 200, phone_bundle.text

        report = client.get(
            "/api/traceability/report.xlsx", params={"by": "credit", "query": "CRED-001"}
        )
        assert report.status_code == 200, report.text
        workbook = load_workbook(io.BytesIO(report.content), read_only=True)
        headers = [cell.value for cell in next(workbook["CDRs"].iter_rows())]
        credit_column = headers.index("Credito") + 1
        trunk_column = headers.index("Troncal de salida") + 1
        trunk_id_column = headers.index("ID troncal de salida") + 1
        assert {workbook["CDRs"].cell(row, credit_column).value for row in (2, 3)} == {"CRED-001"}
        assert {workbook["CDRs"].cell(row, trunk_column).value for row in (2, 3)} == {
            "Troncal principal"
        }
        assert {workbook["CDRs"].cell(row, trunk_id_column).value for row in (2, 3)} == {
            "default"
        }
        workbook.close()

        english_report = client.get(
            "/api/traceability/report.xlsx",
            params={"by": "credit", "query": "CRED-001", "lang": "en"},
        )
        english_workbook = load_workbook(io.BytesIO(english_report.content), read_only=True)
        assert "Overview" in english_workbook.sheetnames
        english_workbook.close()

        bundle = client.get(
            "/api/traceability/bundle.zip", params={"by": "credit", "query": "CRED-001"}
        )
        assert bundle.status_code == 200, bundle.text
        with ZipFile(io.BytesIO(bundle.content)) as archive:
            names = archive.namelist()
            assert "reporte-trazabilidad.xlsx" in names
            assert "manifest-grabaciones.csv" in names
            recordings = [name for name in names if name.startswith("grabaciones/")]
            assert recordings == [f"grabaciones/{first['id']}.ogg"]
            manifest = archive.read("manifest-grabaciones.csv").decode("utf-8-sig")
            assert "troncal,id_troncal" in manifest
            assert "Troncal principal,default" in manifest
            assert "sin_grabacion" in manifest and "CRED-001" in manifest

        english_bundle = client.get(
            "/api/traceability/bundle.zip",
            params={"by": "credit", "query": "CRED-001", "lang": "en"},
        )
        with ZipFile(io.BytesIO(english_bundle.content)) as archive:
            assert "tracking-report.xlsx" in archive.namelist()
            assert "recording-manifest.csv" in archive.namelist()
        with sqlite3.connect(app.state.analytics.path) as db:
            actions = {
                row[0]
                for row in db.execute("SELECT action FROM audit WHERE action LIKE 'traceability.%'")
            }
        assert actions == {
            "traceability.report_downloaded", "traceability.bundle_downloaded"
        }


def test_trace_report_refuses_partial_export(tmp_path):
    settings = Settings(
        data_dir=tmp_path, mode="simulation", report_max_rows=100, auth={"enabled": False}
    )
    settings.report_max_rows = 1
    app = create_app(settings)
    with TestClient(app) as client:
        for name, phone in (("Uno", "5512345678"), ("Dos", "5587654321")):
            client.post("/api/campaigns", json=payload(name, "CRED-LIMIT", phone))
        finish_all(app)
        response = client.get(
            "/api/traceability/report.xlsx",
            params={"by": "credit", "query": "CRED-LIMIT"},
        )
        assert response.status_code == 422 and "supera" in response.json()["detail"]


def test_bundle_rejects_ready_recording_with_untrusted_filename(tmp_path):
    # This protects mass export even if a database file was altered outside the application.
    app = create_app(Settings(data_dir=tmp_path, auth={"enabled": False}))
    with TestClient(app) as client:
        response = client.post(
            "/api/campaigns", json=payload("Uno", "CRED-SAFE", "5512345678")
        )
        cid = response.json()["id"]
        finish_all(app)
        job = client.get(f"/api/campaigns/{cid}/jobs").json()[0]
        with sqlite3.connect(app.state.analytics.path) as db:
            db.execute(
                "INSERT INTO recordings(job_id,status,filename,started_at,evidence) "
                "VALUES(?,'ready','../../otro.ogg',?,'test')",
                (job["id"], job["started_at"]),
            )
        response = client.get(
            "/api/traceability/bundle.zip", params={"by": "credit", "query": "CRED-SAFE"}
        )
        with ZipFile(io.BytesIO(response.content)) as archive:
            assert not [name for name in archive.namelist() if name.startswith("grabaciones/")]
