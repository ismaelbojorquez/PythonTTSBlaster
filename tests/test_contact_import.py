import io
from datetime import date

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from blaster.config import Settings
from blaster.contact_files import import_contacts, read_csv
from blaster.models import parse_contacts, render_message
from blaster.tts import write_tone
from blaster.web import create_app


def workbook_bytes(rows, extra_sheet=False):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contactos"
    for row in rows:
        sheet.append(row)
    if extra_sheet:
        other = workbook.create_sheet("Otro envío")
        other.append(["Credito", "Teléfono", "Cliente", "Producto"])
        other.append(["CRED-3", 5550000103, "Sofía", "Servicio anual"])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


@pytest.mark.parametrize("separator", [",", ";", "\t"])
def test_custom_headers_accents_spaces_and_values_are_literal(separator):
    content = separator.join(
        ["Credito", " Teléfono ", "Nombre completo", "Saldo ($)", "Razón social"]
    )
    content += "\n" + separator.join(
        ["CRED-1", "5550000101", "Ana + Luis", "1250.50", "{{No ejecutar}}"]
    )
    table = read_csv(content)
    contact = parse_contacts(table.csv_text(), "MX")[0]
    assert contact.phone == "525550000101"
    assert (
        render_message(
            "{Nombre completo}: {Saldo ($)} de {Razón social}. {Teléfono}", contact.variables
        )
        == "Ana + Luis: 1250.50 de {{No ejecutar}}. 525550000101"
    )
    assert contact.credit_id == "CRED-1"
    assert table.metadata()["variables"][2] == {
        "name": "Nombre completo", "sample": "Ana + Luis"
    }


def test_literal_keys_never_evaluate_objects_or_reexpand_values():
    assert (
        render_message(
            "{a.b} {x[0]} {Saldo: MXN}",
            {
                "a.b": "{otra}",
                "x[0]": "literal",
                "Saldo: MXN": "100",
            },
        )
        == "{otra} literal 100"
    )
    assert render_message("{{{Nombre}}}", {"Nombre": "Ana"}) == "{Ana}"
    with pytest.raises(ValueError, match="Falta la variable"):
        render_message("{cliente.__class__}", {"cliente": "Ana"})


def test_xlsx_retains_dates_numeric_phones_empty_cells_and_sheet_choice():
    data = workbook_bytes(
        [
            ["credito", "telefono", "Cliente", "Fecha de pago", "Importe", "Nota"],
            ["CRED-1", 5550000101, "Ana", date(2026, 9, 10), 2500.5, None],
            ["CRED-2", 5550000102, "Luis", date(2026, 9, 11), 700, "Otro dato"],
        ],
        extra_sheet=True,
    )
    result = import_contacts(data, "contactos.xlsx")
    assert result["sheet"] == "Contactos" and result["count"] == 2
    contacts = parse_contacts(result["csv_text"], "MX")
    assert contacts[0].variables == {
        "Cliente": "Ana",
        "Fecha de pago": "2026-09-10",
        "Importe": "2500.5",
        "Nota": "",
    }
    other = import_contacts(data, "contactos.xlsx", "Otro envío")
    assert other["count"] == 1
    assert parse_contacts(other["csv_text"], "MX")[0].variables["Producto"] == "Servicio anual"


def test_xlsx_skips_cover_sheet_and_preserves_zero_padded_numbers():
    workbook = Workbook()
    workbook.active.append(["Instrucciones"])
    sheet = workbook.create_sheet("Datos")
    sheet.append(["credito", "telefono", "Folio"])
    sheet.append([42, 2079460018, 42])
    sheet["A2"].number_format = "000000"
    sheet["B2"].number_format = "00000000000"
    sheet["C2"].number_format = "00000"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    result = import_contacts(output.getvalue(), "datos.xlsx")
    assert result["sheet"] == "Datos"
    contact = parse_contacts(result["csv_text"], "GB")[0]
    assert contact.credit_id == "000042"
    assert contact.phone == "442079460018" and contact.variables["Folio"] == "00042"


@pytest.mark.parametrize(
    "headers",
    [
        ["credito", "telefono", "Nombre", " Nombre "],
        ["credito", "telefono", ""],
        ["credito", "telefono", "{Nombre}"],
        ["credito", "telefono", "x" * 65],
        ["credito", "telefono", "Teléfono"],
        ["credito", "Crédito", "telefono"],
        ["credito", "telefono", *[f"col{i}" for i in range(101)]],
    ],
)
def test_invalid_headers_are_rejected_instead_of_losing_columns(headers):
    with pytest.raises(ValueError):
        import_contacts(workbook_bytes([headers, ["5550000101"] * len(headers)]), "datos.xlsx")


def test_xlsx_rejects_formulas_corruption_and_data_without_headers():
    for data in [
        b"not an excel",
        workbook_bytes(
            [["credito", "telefono", "Total"], ["CRED-1", 5550000101, "=1+2"]]
        ),
        workbook_bytes(
            [["credito", "telefono"], ["CRED-1", 5550000101, "dato sin encabezado"]]
        ),
        workbook_bytes(
            [["credito", "telefono"], ["CRED-1", 5550000101] + [None] * 110 + ["dato lejano"]]
        ),
    ]:
        with pytest.raises(ValueError):
            import_contacts(data, "datos.xlsx")


def test_csv_encodings_and_more_than_thirty_custom_variables():
    content = "credito;telefono;Razón social\nCRED-1;5550000101;Compañía"
    for encoding in ["utf-8-sig", "cp1252", "utf-16"]:
        result = import_contacts(content.encode(encoding), "datos.csv")
        assert parse_contacts(result["csv_text"], "MX")[0].variables["Razón social"] == "Compañía"
    headers = ["credito", "telefono", *[f"Campo {i}" for i in range(100)]]
    result = import_contacts(
        workbook_bytes([headers, ["CRED-1", 5550000101] + ["Dato"] * 100]),
        "datos.xlsx",
    )
    contact = parse_contacts(result["csv_text"], "MX")[0]
    assert len(contact.variables) == 100
    assert render_message("{Campo 99}", contact.variables) == "Dato"


def test_import_preview_audio_and_database_share_each_rows_variables(tmp_path):
    app = create_app(Settings(data_dir=tmp_path, auth={"enabled": False}))
    spoken = []

    class Voice:
        async def synthesize(self, text, path):
            spoken.append(text)
            write_tone(path, 0.05)

    with TestClient(app) as client:
        app.state.speech_preview.speech = Voice()
        data = workbook_bytes(
            [
                ["Credito", "telefono", "Nombre completo", "Pago pendiente"],
                ["CRED-1", 5550000101, "Ana", 1250],
                ["CRED-2", 5550000102, "Luis", 850],
            ]
        )
        response = client.post(
            "/api/contacts/import?filename=datos.xlsx",
            content=data,
            headers={"Content-Type": "application/octet-stream"},
        )
        assert response.status_code == 200, response.text
        csv_text = response.json()["csv_text"]
        inspected = client.post("/api/contacts/inspect", json={"csv_text": csv_text})
        assert inspected.json()["variables"][3]["name"] == "Pago pendiente"
        payload = {
            "name": "Variables libres",
            "agent_number": "5550000103",
            "template": "Hola {Nombre completo}, tu pago es {Pago pendiente}.",
            "csv_text": csv_text,
            "country": "MX",
        }
        preview = client.post("/api/preview", json=payload)
        assert preview.status_code == 200, preview.text
        assert [s["message"] for s in preview.json()["samples"]] == [
            "Hola Ana, tu pago es 1250.",
            "Hola Luis, tu pago es 850.",
        ]
        audio = client.post(
            "/api/preview/audio",
            json={
                k: payload[k]
                for k in (
                    "template",
                    "csv_text",
                    "country",
                )
            },
        )
        assert audio.status_code == 200 and spoken[0].startswith("Hola Ana, tu pago es 1250.")
        campaign = client.post("/api/campaigns", json=payload)
        assert campaign.status_code == 201, campaign.text
        jobs = client.get(f"/api/campaigns/{campaign.json()['id']}/jobs").json()
        assert jobs[1]["credit_id"] == "CRED-2"
        assert jobs[1]["variables"]["Pago pendiente"] == "850"
        assert jobs[1]["message"] == "Hola Luis, tu pago es 850."
        missing = client.post("/api/campaigns", json={**payload, "template": "Hola {otra columna}"})
        assert missing.status_code == 422 and "otra columna" in missing.json()["detail"]
        assert client.get("/api/status").json()["active_sessions"] == 0

        saved = client.post("/api/manage/templates", json={
            "name": "Encabezados literales",
            "message": "Hola {Nombre completo}, {Saldo: MXN}, {cliente.id}, {A[1]}.",
        })
        assert saved.status_code == 200, saved.text
        assert client.get("/api/manage/templates").json()[0]["message"].endswith("{A[1]}.")


def test_import_requires_session_and_respects_origin_and_size(tmp_path):
    app = create_app(Settings(data_dir=tmp_path))
    with TestClient(app) as client:
        headers = {"Content-Type": "application/octet-stream"}
        url = "/api/contacts/import?filename=datos.csv"
        assert client.post(url, content=b"telefono\n5550000101", headers=headers).status_code == 401
        assert (
            client.post(
                url,
                content=b"x",
                headers={
                    **headers,
                    "Origin": "https://untrusted.example",
                },
            ).status_code
            == 403
        )
        assert client.post(url, content=b"x" * 8_500_001, headers=headers).status_code == 413
