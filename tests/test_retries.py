import asyncio
import csv
import io
import json
import sqlite3
from datetime import datetime, timedelta
from uuid import uuid4

import pytest
from amd_samples import signal, silence
from conftest import remove_retries_schema
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from test_campaign_execution import CAMPAIGN, make_app
from test_engine import until
from test_management import app_for, setup

from blaster.analytics import Analytics, Filters
from blaster.models import CampaignInput, Contact
from blaster.reports import cdr_csv, excel_report
from blaster.retries import RetryPolicy
from blaster.store import Store, now


def create(store, **policy):
    return store.create_campaign(CampaignInput(
        name="Reintentos", template="Hola {nombre}", agent_number="525550009999",
        contacts=[Contact(
            phone="525550000100", credit_id="CRED-ANA", variables={"nombre": "Ana"}
        )],
        retry_policy=RetryPolicy(max_attempts=policy.pop("max_attempts", 3),
                                 delay_seconds=policy.pop("delay_seconds", 1), **policy),
    ))


@pytest.mark.parametrize("code,outcome", [
    (480, "no_answer"), (486, "busy"),
    (503, "temporary_error"), (504, "temporary_error"),
])
async def test_retry_waits_without_channels_and_keeps_every_cdr(engine, code, outcome):
    cid = create(engine.store, max_attempts=2,
                 outcomes=["no_answer", "busy", "temporary_error"])
    original = engine.store.jobs(cid)[0]
    engine.telephony.outcomes[original["phone"]] = code
    engine.start_campaign(cid)
    await until(lambda: engine.store.jobs(cid)[0]["attempt_number"] == 2)
    pending = engine.store.jobs(cid)[0]
    first = engine.store.jobs(cid, latest=False)[0]
    assert first["status"] == outcome and pending["status"] == "queued"
    assert first["id"] != pending["id"] and pending["retry_of"] == first["id"]
    assert pending["contact_id"] == first["id"]
    assert engine.snapshot()["reserved_channels"] == 0 and not engine.sessions
    assert engine.active_campaign == cid and engine.store.next_queued(cid) is None
    assert engine.store.campaigns()[0]["total"] == 1
    assert engine.store.retries.plan(first["id"]) == pending["id"]
    assert len(engine.store.jobs(cid, latest=False)) == 2
    if code in {503, 504}:
        # Routing has its own backoff, independent of the contact retry timer.
        engine.router.cooldown.clear()
    await until(lambda: engine.active_campaign is None, timeout=5)
    records = Analytics(engine.settings.data_dir / "test.sqlite3")
    rows, summary, events = records.report_data(Filters(mode="simulation"), 100)
    assert len(rows) == 2 and {r["attempt_number"] for r in rows} == {1, 2}
    first_cdr = records.detail(first["id"])
    second_cdr = records.detail(pending["id"])
    assert datetime.fromisoformat(second_cdr["started_at"]) >= (
        datetime.fromisoformat(first_cdr["finalized_at"]) + timedelta(seconds=1)
    )
    assert first_cdr["retry_decision"]["next_job_id"] == pending["id"]
    assert second_cdr["retry_decision"]["reason"] == "attempt_limit"
    assert len(first_cdr["attempts"]) == 2
    assert len(list(csv.DictReader(io.StringIO(cdr_csv(rows).decode("utf-8-sig"))))) == 2
    book = load_workbook(io.BytesIO(
        excel_report(rows, summary, events, Filters(mode="simulation"))
    ))
    assert book["CDRs"].max_row == 3
    headers = [cell.value for cell in book["CDRs"][1]]
    column = headers.index("Intento") + 1
    assert {book["CDRs"].cell(i, column).value for i in (2, 3)} == {1, 2}


@pytest.mark.parametrize("audio,status", [
    (signal(3000), "machine"), (silence(3000), "amd_unknown"),
])
async def test_amd_retry_stops_after_human_answer(engine, audio, status):
    engine.settings.amd.enabled = True
    engine.settings.amd.unknown_action = "hangup"
    cid = create(engine.store)
    job = engine.store.jobs(cid)[0]
    engine.telephony.amd_audio[job["phone"]] = audio
    engine.start_campaign(cid)
    await until(lambda: engine.store.jobs(cid)[0]["attempt_number"] == 2)
    assert engine.store.jobs(cid, latest=False)[0]["status"] == status
    engine.telephony.amd_audio.pop(job["phone"])
    await until(lambda: engine.active_campaign is None, timeout=5)
    rows = engine.store.jobs(cid, latest=False)
    assert len(rows) == 2 and rows[1]["status"] == "no_input"
    detail = Analytics(engine.settings.data_dir / "test.sqlite3").detail(rows[1]["id"])
    assert detail["amd_verdict"] == "human"
    assert detail["retry_decision"]["reason"] == "contact_reached"


@pytest.mark.parametrize("code", [403, 404, 484, 603])
async def test_permanent_sip_errors_do_not_retry(engine, code):
    cid = create(engine.store, outcomes=["temporary_error"])
    engine.telephony.outcomes[engine.store.jobs(cid)[0]["phone"]] = code
    engine.start_campaign(cid)
    await until(lambda: engine.active_campaign is None)
    assert len(engine.store.jobs(cid, latest=False)) == 1


async def test_disabled_amd_message_played_never_retries_and_busy_may_be_excluded(engine):
    cid = create(engine.store)
    engine.start_campaign(cid)
    await until(lambda: engine.active_campaign is None)
    job = engine.store.jobs(cid)[0]
    assert job["status"] == "no_input" and job["attempt_number"] == 1
    cid = create(engine.store, outcomes=["machine"])
    engine.telephony.outcomes[job["phone"]] = 486
    engine.start_campaign(cid)
    await until(lambda: engine.active_campaign is None)
    assert len(engine.store.jobs(cid, latest=False)) == 1


async def test_pause_resume_and_stop_pending_retry(engine):
    cid = create(engine.store)
    engine.telephony.outcomes[engine.store.jobs(cid)[0]["phone"]] = 486
    engine.start_campaign(cid)
    await until(lambda: engine.store.jobs(cid)[0]["attempt_number"] == 2)
    engine.pause_campaign(cid)
    await asyncio.sleep(1.15)
    assert engine.store.jobs(cid)[0]["status"] == "queued"
    engine.start_campaign(cid)
    await until(lambda: engine.store.jobs(cid)[0]["attempt_number"] == 3)
    await engine.stop_campaign(cid)
    assert engine.store.jobs(cid)[0]["status"] == "cancelled"
    assert not engine.store.has_queued(cid) and not engine.sessions
    engine.store.retries.reconcile(cid)
    assert len(engine.store.jobs(cid, latest=False)) == 3


def finished(store, cid, status="busy", **fields):
    jid = store.jobs(cid)[0]["id"]
    store.set_campaign_status(cid, "running")
    store.begin_call(jid, True)
    store.transition(jid, "dialing")
    store.transition(jid, status)
    store.update_call(jid, finalized_at=now(), **fields)
    return jid


def test_recovery_plans_once_and_preserves_deadline_and_copy_policy(tmp_path):
    path = tmp_path / "test.sqlite3"
    store = Store(path)
    cid = create(store, delay_seconds=300)
    first = finished(store, cid)  # Simulate crash after CDR finalization, before retry insertion.
    store.close()
    store = Store(path)
    store.recover()
    second = store.jobs(cid)[0]
    assert second["attempt_number"] == 2 and second["retry_of"] == first
    assert store.campaign(cid)["status"] == "paused"
    store.recover()
    assert store.jobs(cid)[0] == second
    store.close()
    store = Store(path)
    store.recover()
    assert store.jobs(cid)[0] == second and not store.next_queued(cid)
    duplicate = store.history.copy(cid, "duplicate", "Copia", "", str(uuid4()), {},
                                   "simulation", "fingerprint")
    copied = store.jobs(duplicate)
    assert len(copied) == 1 and copied[0]["attempt_number"] == 1
    assert copied[0]["contact_id"] == copied[0]["id"] != first
    assert copied[0]["available_at"] is None and copied[0]["retry_of"] is None
    assert store.campaign(cid)["retry_policy"] == store.campaign(duplicate)["retry_policy"]
    assert store.history.history(cid)["items"][0]["contacts"] == 1
    store.close()


def test_retry_transaction_rolls_back_when_audit_fails(tmp_path):
    store = Store(tmp_path / "test.sqlite3")
    cid = create(store)
    jid = finished(store, cid)
    store.db.execute("CREATE TRIGGER reject_audit BEFORE INSERT ON audit "
                     "BEGIN SELECT RAISE(ABORT,'audit unavailable'); END")
    with pytest.raises(sqlite3.IntegrityError):
        store.retries.plan(jid)
    assert len(store.jobs(cid, latest=False)) == 1
    assert not store.db.execute("SELECT 1 FROM retry_decisions").fetchone()
    store.close()


def test_retry_requires_confirmed_disconnect_and_avoids_interrupted_calls(tmp_path):
    store = Store(tmp_path / "test.sqlite3")
    cid = create(store)
    jid = finished(store, cid, "no_answer")
    with store.db:
        store.db.execute(
            "INSERT INTO call_legs(id,job_id,role,number,created_at,invite_at) "
            "VALUES(? ,?,'customer','525550000100',?,?)", (uuid4().hex, jid, now(), now()),
        )
    assert store.retries.plan(jid) is None
    assert store.db.execute("SELECT reason FROM retry_decisions WHERE job_id=?", (jid,)
                            ).fetchone()[0] == "unconfirmed_disconnect"
    cid = create(store)
    jid = finished(store, cid, "amd_unknown", end_reason="process_interrupted")
    assert store.retries.plan(jid) is None
    assert len(store.jobs(cid, latest=False)) == 1
    store.close()


def test_policy_validation_and_draft_only_edit_with_audit(tmp_path):
    app = make_app(tmp_path)
    with TestClient(app) as client:
        policy = {"max_attempts": 3, "delay_seconds": 600, "outcomes": ["no_answer", "machine"]}
        cid = client.post("/api/campaigns", json={**CAMPAIGN, "retry_policy": policy}).json()["id"]
        assert client.get("/api/campaigns").json()[0]["retry_policy"] == policy
        for invalid in ({"max_attempts": 11}, {"delay_seconds": 0},
                        {"max_attempts": 2, "outcomes": []}, {"outcomes": ["completed"]}):
            assert client.post(f"/api/campaigns/{cid}/retries", json=invalid).status_code == 422
        changed = {**policy, "delay_seconds": 120}
        assert client.post(f"/api/campaigns/{cid}/retries", json=changed).status_code == 200
        with sqlite3.connect(tmp_path / "blaster.sqlite3") as db:
            audit = json.loads(db.execute("SELECT detail FROM audit WHERE "
                                         "action='campaign.retries_updated'").fetchone()[0])
            assert audit == {"before": policy, "after": changed}
        client.post(f"/api/campaigns/{cid}/start", json={})
        assert client.post(f"/api/campaigns/{cid}/retries", json=policy).status_code == 422
        client.post(f"/api/campaigns/{cid}/stop", json={})


def test_retry_edit_requires_operator_and_valid_origin(tmp_path):
    with TestClient(app_for(tmp_path)) as client:
        setup(client)
        cid = client.post("/api/campaigns", json=CAMPAIGN).json()["id"]
        analyst = {"username": "analista", "password": "long-test-password",
                   "display_name": "Analista de prueba", "role": "analyst"}
        assert client.post("/api/manage/users", json=analyst).status_code == 200
        assert client.post(f"/api/campaigns/{cid}/retries", json={},
                           headers={"Origin": "https://other.invalid"}).status_code == 403
        client.post("/api/auth/logout", json={})
        assert client.post("/api/auth/login", json=analyst).status_code == 200
        assert client.post(f"/api/campaigns/{cid}/retries", json={}).status_code == 403


def test_v5_upgrade_defaults_to_single_attempt_and_keeps_history(tmp_path):
    path = tmp_path / "v5.sqlite3"
    store = Store(path)
    cid = create(store)
    first = finished(store, cid)
    with store.db:
        remove_retries_schema(store.db)
        store.db.execute("PRAGMA user_version=5")
    store.close()
    for _ in range(2):
        store = Store(path)
        store.recover()
        assert store.campaign(cid)["retry_policy"]["max_attempts"] == 1
        jobs = store.jobs(cid)
        assert len(jobs) == 1 and jobs[0]["id"] == jobs[0]["contact_id"] == first
        assert jobs[0]["status"] == "busy"
        store.close()
    assert len(list(tmp_path.glob("*.before-retries-*.bak"))) == 1
