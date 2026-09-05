"""Contact tables and literal column names shared by import and campaign validation."""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

MAX_BYTES = 8_000_000
MAX_COLUMNS = 102  # Credit, telephone and up to 100 user-defined variables.


def column_name(value: str) -> str:
    name = unicodedata.normalize("NFC", value.strip())
    if not name or len(name) > 64 or any(c in name for c in "{}\r\n\t"):
        raise ValueError("Cada encabezado debe tener de 1 a 64 caracteres, sin llaves ni saltos")
    return name


def phone_column(headers: list[str]) -> str:
    aliases = {"telefono", "teléfono", "phone", "telephone"}
    matches = [h for h in headers if h.casefold() in aliases]
    if len(matches) != 1:
        raise ValueError(
            "Incluye una sola columna Telefono/Phone "
            "(también se admiten Teléfono y Telephone)"
        )
    return matches[0]


def credit_column(headers: list[str]) -> str:
    aliases = {"credito", "crédito", "credit", "account", "account_id"}
    matches = [h for h in headers if h.casefold() in aliases]
    if len(matches) != 1:
        raise ValueError(
            "Incluye una sola columna Credito/Account "
            "(también se admiten Crédito, Credit y Account ID)"
        )
    return matches[0]


@dataclass
class ContactTable:
    headers: list[str]
    rows: list[list[str]]

    def csv_text(self) -> str:
        output = io.StringIO(newline="")
        writer = csv.writer(output, lineterminator="\n")
        writer.writerow(self.headers)
        writer.writerows(self.rows)
        result = output.getvalue()
        if len(result.encode("utf-8")) > MAX_BYTES:
            raise ValueError("Los contactos superan 8 MB. Divide la lista en campañas más pequeñas")
        return result

    def metadata(self) -> dict:
        return {
            "count": len(self.rows),
            "variables": [
                {"name": name, "sample": self.rows[0][i] if self.rows else ""}
                for i, name in enumerate(self.headers)
            ],
        }


def table_from_rows(rows) -> ContactTable:
    iterator = iter(rows)
    try:
        raw_headers = next(iterator)
    except StopIteration as error:
        raise ValueError("El archivo no contiene encabezados") from error
    if len(raw_headers) > MAX_COLUMNS:
        raise ValueError("Máximo 100 variables además de Credito y Telefono")
    headers = [column_name(value) for value in raw_headers]
    if len(set(headers)) != len(headers):
        raise ValueError("El archivo contiene columnas repetidas")
    phone_column(headers)
    credit_column(headers)
    result = []
    for line, row in enumerate(iterator, 2):
        if not row or not any(value.strip() for value in row):
            continue
        if len(row) != len(headers):
            raise ValueError(f"La fila {line} tiene columnas incompletas o extra")
        if any(len(value) > 1000 for value in row):
            raise ValueError(f"Fila {line}: máximo 1000 caracteres por celda")
        if len(result) >= 10000:
            raise ValueError("Máximo 10 000 contactos por campaña")
        result.append(row)
    if not result:
        raise ValueError("El archivo no contiene contactos")
    return ContactTable(headers, result)


def read_csv(content: str) -> ContactTable:
    if len(content.encode("utf-8")) > MAX_BYTES:
        raise ValueError("El archivo supera 8 MB")
    content = content.lstrip("\ufeff")
    try:
        dialect = csv.Sniffer().sniff(content[:8192], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    try:
        return table_from_rows(csv.reader(io.StringIO(content, newline=""), dialect, strict=True))
    except csv.Error as error:
        raise ValueError(
            "CSV inválido: revisa las comillas y los separadores de las columnas"
        ) from error


def cell_text(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if cell.data_type in {"f", "e"}:
        raise ValueError(
            f"Celda {cell.coordinate}: reemplaza las fórmulas o errores "
            "por valores antes de importar"
        )
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time() == time() else value.isoformat(sep=" ")
    if isinstance(value, date | time):
        return value.isoformat()
    if isinstance(value, bool):
        return "Sí" if value else "No"
    if isinstance(value, int | float):
        text = str(int(value)) if int(value) == value else str(value)
        if value >= 0 and int(value) == value and re.fullmatch(r"0{1,64}", cell.number_format):
            text = text.zfill(len(cell.number_format))
        return text
    return str(value)


def read_xlsx(data: bytes, sheet: str = "") -> tuple[ContactTable, list[str], str]:
    try:
        with ZipFile(io.BytesIO(data)) as archive:
            if (
                len(archive.infolist()) > 2000
                or sum(item.file_size for item in archive.infolist()) > 64_000_000
            ):
                raise ValueError(
                    "El Excel es demasiado grande al descomprimirlo; divide el archivo"
                )
        workbook = load_workbook(
            io.BytesIO(data), read_only=True, data_only=False, keep_links=False
        )
    except (BadZipFile, InvalidFileException, KeyError, OSError, SyntaxError) as error:
        raise ValueError("No se pudo abrir el Excel. Guarda un archivo .xlsx válido") from error
    try:
        sheets = [ws.title for ws in workbook.worksheets if ws.sheet_state == "visible"]
        if not sheets:
            raise ValueError("El Excel no contiene hojas visibles")
        if sheet and sheet not in sheets:
            raise ValueError("Selecciona una hoja visible del Excel")
        first_error = None
        for selected in [sheet] if sheet else sheets:
            try:
                table = worksheet_table(workbook[selected])
                return table, sheets, selected
            except ValueError as error:
                if first_error is None:
                    first_error = error
        raise first_error
    finally:
        workbook.close()


def worksheet_table(worksheet) -> ContactTable:
    # Some exporters store incorrect dimensions. Read actual rows with an explicit limit.
    worksheet.reset_dimensions()
    rows = worksheet.iter_rows()
    header = [cell_text(c) for c in next(rows, [])]
    while header and not header[-1].strip():
        header.pop()
    width = len(header)

    def values():
        yield header
        for index, row in enumerate(rows, 2):
            if index > 10001:
                raise ValueError(
                    "Máximo 10 000 filas de contactos en cada hoja; elimina filas vacías"
                )
            texts = [cell_text(c) for c in row]
            if any(texts[width:]):
                raise ValueError("El Excel contiene datos en columnas sin encabezado")
            yield texts[:width] + [""] * max(0, width - len(texts))

    return table_from_rows(values())


def import_contacts(data: bytes, filename: str, sheet: str = "") -> dict:
    if len(data) > MAX_BYTES:
        raise ValueError("El archivo supera 8 MB")
    suffix = Path(filename).suffix.lower()
    sheets, selected = [], ""
    if suffix == ".xlsx":
        try:
            table, sheets, selected = read_xlsx(data, sheet)
        except ValueError:
            raise
        except Exception as error:
            raise ValueError("No se pudo leer el Excel. Guarda un archivo .xlsx válido") from error
    elif suffix == ".csv":
        try:
            content = data.decode(
                "utf-16" if data.startswith((b"\xff\xfe", b"\xfe\xff")) else "utf-8-sig"
            )
        except UnicodeDecodeError:
            try:
                content = data.decode("cp1252")
            except UnicodeDecodeError as error:
                raise ValueError(
                    "No se pudo leer el CSV. Guárdalo con codificación UTF-8"
                ) from error
        table = read_csv(content)
    else:
        raise ValueError("Selecciona un archivo CSV o XLSX")
    return {**table.metadata(), "csv_text": table.csv_text(), "sheets": sheets, "sheet": selected}
