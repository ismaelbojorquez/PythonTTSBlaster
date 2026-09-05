"""Excel and CDR exports generated inside Python, without a reporting service."""

from __future__ import annotations

import csv
import io
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

from blaster.analytics import ACTOR_LABELS, AMD_LABELS, LEG_FIELDS, STATUS_LABELS

CDR_COLUMNS = [
    ("id", "ID llamada"),
    ("campaign_name", "Campaña"),
    ("campaign_id", "ID campaña"),
    ("mode", "Modo"),
    ("contact_name", "Nombre en contacto"),
    ("credit_id", "Credito"),
    ("phone", "Teléfono cliente"),
    ("customer_trunk_name", "Troncal de salida"),
    ("customer_trunk_id", "ID troncal de salida"),
    ("agent_number", "Número agente"),
    ("agent_trunk_name", "Troncal de transferencia"),
    ("agent_trunk_id", "ID troncal de transferencia"),
    ("status_label", "Resultado"),
    ("coverage", "Cobertura"),
    ("started_at", "Inicio"),
    ("customer_invite_at", "INVITE cliente"),
    ("customer_ringing_at", "Timbrado cliente"),
    ("customer_answered_at", "Respuesta cliente"),
    ("customer_media_at", "Audio cliente activo"),
    ("customer_ended_at", "Fin tramo cliente"),
    ("customer_pdd_seconds", "Demora hasta timbrado (s)"),
    ("customer_setup_seconds", "Demora hasta respuesta (s)"),
    ("customer_connected_seconds", "Cliente conectado (s)"),
    ("customer_total_seconds", "Tramo cliente total (s)"),
    ("amd_label", "Resultado AMD"),
    ("amd_reason", "Motivo AMD"),
    ("amd_elapsed_ms", "Análisis AMD (ms)"),
    ("amd_voiced_ms", "Voz AMD (ms)"),
    ("amd_words", "Segmentos de voz AMD"),
    ("tts_ms", "Generación TTS (ms)"),
    ("message_started_at", "Inicio reproducción"),
    ("message_completed_at", "Mensaje completo"),
    ("replays", "Repeticiones"),
    ("transfer_requested_at", "Solicitud agente"),
    ("transfer_actor", "Quién solicitó agente"),
    ("agent_invite_at", "INVITE agente"),
    ("agent_ringing_at", "Timbrado agente"),
    ("agent_answered_at", "Respuesta agente"),
    ("agent_connected_seconds", "Agente conectado (s)"),
    ("bridged_at", "Inicio puente"),
    ("bridge_ended_at", "Fin puente"),
    ("bridge_seconds", "Conversación en puente (s)"),
    ("end_actor_label", "Quién inició fin sesión"),
    ("end_reason", "Motivo fin"),
    ("end_evidence", "Evidencia fin"),
    ("customer_sip_code", "SIP cliente"),
    ("agent_sip_code", "SIP agente"),
    ("customer_call_id", "SIP Call-ID cliente"),
    ("agent_call_id", "SIP Call-ID agente"),
    ("finalized_at", "Registro finalizado"),
    ("detail", "Detalle"),
    ("agent_strategy", "Distribución de transferencias"),
    ("agent_pool_wait_seconds", "Espera de teléfono libre (s)"),
    ("contact_id", "ID contacto en campaña"),
    ("attempt_number", "Intento"),
    ("retry_of", "ID intento anterior"),
    ("available_at", "Reintento disponible desde"),
]
CDR_COLUMNS_EN = {
    "ID llamada": "Call ID", "Campaña": "Campaign", "ID campaña": "Campaign ID",
    "Modo": "Operation type", "Nombre en contacto": "Contact name", "Credito": "Account",
    "Teléfono cliente": "Customer phone", "Troncal de salida": "Outbound provider",
    "ID troncal de salida": "Outbound provider ID", "Número agente": "Agent phone",
    "Troncal de transferencia": "Transfer provider", "ID troncal de transferencia": "Transfer provider ID",
    "Resultado": "Outcome", "Cobertura": "Data coverage", "Inicio": "Started",
    "INVITE cliente": "Customer request sent", "Timbrado cliente": "Customer ringing",
    "Respuesta cliente": "Customer answered", "Audio cliente activo": "Customer audio active",
    "Fin tramo cliente": "Customer leg ended", "Demora hasta timbrado (s)": "Time to ringing (s)",
    "Demora hasta respuesta (s)": "Time to answer (s)", "Cliente conectado (s)": "Customer connected (s)",
    "Tramo cliente total (s)": "Total customer leg (s)", "Resultado AMD": "Answer classification",
    "Motivo AMD": "Classification reason", "Análisis AMD (ms)": "Answer analysis (ms)",
    "Voz AMD (ms)": "Detected speech (ms)", "Segmentos de voz AMD": "Speech segments",
    "Generación TTS (ms)": "Voice preparation (ms)", "Inicio reproducción": "Message started",
    "Mensaje completo": "Message completed", "Repeticiones": "Replays",
    "Solicitud agente": "Agent requested", "Quién solicitó agente": "Agent requested by",
    "INVITE agente": "Agent request sent", "Timbrado agente": "Agent ringing",
    "Respuesta agente": "Agent answered", "Agente conectado (s)": "Agent connected (s)",
    "Inicio puente": "Conversation connected", "Fin puente": "Conversation ended",
    "Conversación en puente (s)": "Agent conversation (s)", "Quién inició fin sesión": "Call ended by",
    "Motivo fin": "End reason", "Evidencia fin": "End evidence", "SIP cliente": "Customer provider code",
    "SIP agente": "Agent provider code", "SIP Call-ID cliente": "Customer provider reference",
    "SIP Call-ID agente": "Agent provider reference", "Registro finalizado": "Record completed",
    "Detalle": "Details", "Distribución de transferencias": "Transfer distribution",
    "Espera de teléfono libre (s)": "Wait for available phone (s)",
    "ID contacto en campaña": "Campaign contact ID", "Intento": "Attempt",
    "ID intento anterior": "Previous attempt ID", "Reintento disponible desde": "Retry available from",
}

REPORT_VALUES_EN = {
    "Finalizada": "Completed", "Buzón probable": "Probable voicemail",
    "Respuesta no identificada": "Unidentified answer", "Sin respuesta": "No answer",
    "Sin selección": "No selection", "Ocupada": "Busy", "Fallida": "Failed",
    "Proveedor no disponible": "Provider unavailable", "Cancelada": "Canceled",
    "Interrumpida": "Interrupted", "Marcando": "Dialing", "Identificando respuesta": "Identifying answer",
    "Preparando mensaje": "Preparing message", "Mensaje en curso": "Message playing",
    "Esperando respuesta": "Waiting for response", "Contactando al agente": "Contacting agent",
    "Esperando agente libre": "Waiting for available agent", "Con agente": "With agent",
    "Pendiente": "Pending", "Cliente": "Customer", "Agente": "Agent", "Plataforma": "Platform",
    "Operador": "Operator", "Proveedor": "Provider", "No identificado": "Unidentified",
    "Persona probable": "Probable person", "Sin evaluación": "Not assessed",
    "Sin información anterior": "No previous information", "sip": "Live", "simulation": "Test",
    "measured": "Captured", "legacy": "Historical",
}
DEFINITIONS = [
    ("Fuente", "Base SQLite local; una fila CDR por intento cuya sesión inició; "
     "los reintentos tienen IDs distintos."),
    (
        "Cobertura measured",
        "Telemetría capturada por esta versión. Una celda vacía significa sin evidencia.",
    ),
    (
        "Cobertura legacy",
        "Registro anterior: se conserva su estado y fechas; no se infieren respuesta ni cuelgue.",
    ),
    (
        "Nombre en contacto",
        "Nombre importado. No identifica ni verifica a la persona que contestó.",
    ),
    (
        "Credito",
        "Identificador obligatorio importado con el contacto; se conserva en cada intento.",
    ),
    (
        "Troncal",
        "Nombre e identificador de la ruta usada para originar cada tramo. La hoja Tramos "
        "conserva todos los intentos cuando existe cambio a una troncal de respaldo.",
    ),
    ("Respuesta", "Respuesta SIP 2xx al INVITE o llamada confirmada; puede ser un buzón."),
    ("AMD", "Clasificación probable por audio; no acredita identidad. Sin IA; puede equivocarse."),
    (
        "Conectado (s)",
        "Desde la respuesta hasta la desconexión observada de cada tramo. Incluye TTS y espera.",
    ),
    (
        "Puente (s)",
        "Desde que se enlaza audio bidireccional hasta la primera desconexión observada.",
    ),
    (
        "ASR",
        "Respuestas observadas / INVITE de cliente observados. Excluye históricos sin telemetría.",
    ),
    ("Éxito transferencia", "Puentes establecidos / solicitudes de agente por DTMF 2."),
    (
        "Promedios",
        "Sólo mediciones disponibles. Tiempos conectados excluyen tramos aún activos "
        "o interrumpidos sin fin.",
    ),
    (
        "Fin remoto",
        "BYE/CANCEL recibido en el tramo cliente/agente. La troncal puede originarlo; "
        "no prueba identidad física.",
    ),
    (
        "Desvíos",
        "Se guardan solicitudes DTMF 2, REFER rechazados y respuestas SIP 3xx observadas. "
        "Un desvío interno del operador puede no ser visible.",
    ),
    (
        "Duraciones",
        "Reloj monotónico, segundos sin redondeo de facturación. No equivalen al CDR "
        "facturado del proveedor.",
    ),
    (
        "Interrupción",
        "Ante cierre abrupto se preserva lo observado y se dejan vacíos los tiempos que "
        "no pudieron medirse.",
    ),
    (
        "Modo",
        "Simulación y SIP se separan mediante el filtro. all mezcla ambos de forma explícita.",
    ),
    (
        "Fechas",
        "Excel usa la zona del reporte; CSV y eventos JSON conservan ISO 8601 UTC con offset.",
    ),
    (
        "Eventos",
        "Cronología técnica sin audio ni cabeceras SIP de autenticación. Los estados "
        "operativos permanecen en SQLite.",
    ),
]
DEFINITIONS_EN = [
    ("Source", "Local SQLite database; one call record per started attempt. Retries have separate IDs."),
    ("Captured coverage", "Tracking captured by this version. A blank cell means no evidence was observed."),
    ("Historical coverage", "Earlier record with its original outcome and dates; answer or hang-up data is not inferred."),
    ("Contact name", "Imported name. It does not identify or verify the person who answered."),
    ("Account", "Required identifier imported with the contact and preserved across every attempt."),
    ("Provider", "Name and identifier of the route used for each call leg. The Call legs sheet keeps every route used during failover."),
    ("Answer", "Confirmed provider answer to the call request; it may be voicemail."),
    ("Answer classification", "Probable audio-based classification. It does not verify identity and may be incorrect."),
    ("Connected time", "From answer to the observed disconnection of each leg. Includes the message and waiting time."),
    ("Agent conversation", "From two-way audio connection until the first observed disconnection."),
    ("Answer rate", "Observed answers divided by observed customer call attempts. Excludes historical records with limited data."),
    ("Transfer success", "Agent conversations divided by agent requests made with option 2."),
    ("Averages", "Only available measurements. Connected times exclude active calls and interrupted legs without a measured ending."),
    ("Remote ending", "A remote ending observed on the customer or agent leg. The provider may originate it; it does not prove physical identity."),
    ("Redirects", "Keeps option 2 requests, rejected redirects, and observed provider redirects. Internal provider redirects may not be visible."),
    ("Durations", "Monotonic elapsed seconds without billing rounding. They may differ from the provider's billed call detail."),
    ("Interruption", "When the application closes abruptly, observed data is preserved and unmeasurable times remain blank."),
    ("Operation type", "Test and live calls are separated by the filter. All combines both explicitly."),
    ("Dates", "Excel uses the report time zone. CSV and event data preserve ISO 8601 timestamps with offsets."),
    ("Events", "Activity timeline without audio or authentication credentials. Operational states remain in the local database."),
]


def _label(value, language: str):
    return REPORT_VALUES_EN.get(value, value) if language == "en" else value


def cdr_csv(rows, language: str = "es"):
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow([CDR_COLUMNS_EN.get(label, label) if language == "en" else label for _, label in CDR_COLUMNS])
    for row in rows:
        values = []
        for key, _ in CDR_COLUMNS:
            value = row.get(key)
            if key in {"status_label", "amd_label", "end_actor_label", "mode", "coverage"}:
                value = _label(value, language)
            if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
                value = "'" + value
            values.append(value)
        writer.writerow(values)
    return output.getvalue().encode("utf-8-sig")


INK, ACCENT, WASH = "203841", "176278", "E3F0F4"


def excel_report(rows, summary, events, filters, language: str = "es"):
    workbook = Workbook(write_only=True)
    workbook.properties.creator = "Blaster TTS"
    english = language == "en"
    tr = lambda es, en: en if english else es
    workbook.properties.title = tr("Analítica de llamadas", "Call analytics")
    zone = ZoneInfo(filters.timezone)

    def sheet(name, headers, widths=None):
        ws = workbook.create_sheet(name)
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
        ws.sheet_properties.outlinePr.summaryRight = False
        from openpyxl.utils import get_column_letter

        for index, label in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(index)].width = (
                widths[index - 1] if widths else min(42, max(23, len(label) + 3))
            )
        append(ws, headers, header=True)
        return ws

    def append(ws, values, *, header=False):
        cells = []
        for value in values:
            cell = WriteOnlyCell(ws, value=value)
            if isinstance(value, str):
                # Force text: user contact/campaign content cannot become Excel formulas.
                cell.data_type = "s"
            cell.font = Font(name="Aptos", size=11, color="FFFFFF" if header else INK, bold=header)
            cell.alignment = Alignment(vertical="top", wrap_text=header)
            if header:
                cell.fill = PatternFill("solid", fgColor=ACCENT)
            elif isinstance(value, (float, int)):
                cell.number_format = "#,##0.00" if isinstance(value, float) else "#,##0"
            elif isinstance(value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            cells.append(cell)
        ws.append(cells)

    def date_value(key, value):
        if key.endswith("_at") and value:
            return datetime.fromisoformat(value).astimezone(zone).replace(tzinfo=None)
        return value

    main = sheet(
        tr("Resumen", "Overview"),
        [tr("BLASTER · Analítica de llamadas", "BLASTER · Call analytics"), tr("Valor", "Value"), tr("Definición", "Definition")],
        [43, 25, 75]
    )
    counts = summary["counts"]
    append(
        main, [tr("Generado", "Generated"), date_value("generated_at", summary["generated_at"]), filters.timezone]
    )
    append(
        main,
        [
            tr("Período", "Period"),
            f"{filters.date_from or tr('Inicio', 'Start')} → {filters.date_to or tr('Actual', 'Current')}",
            f"{tr('Modo', 'Operation type')}: {_label(filters.mode, language)}; {tr('campaña', 'campaign')}: {filters.campaign_id or tr('Todas', 'All')}; "
            + (
                f"{tr('Credito', 'Account')}: {filters.credit_id}"
                if filters.credit_id is not None
                else f"{tr('Telefono', 'Phone')}: {filters.phone}"
                if filters.phone is not None
                else tr("sin identificador exacto", "no exact identifier")
            ),
        ],
    )
    append(main, [tr("Indicador", "Metric"), tr("Resultado", "Result"), tr("Base de cálculo", "Calculation basis")], header=True)
    for key, label, definition in [
        ("total", tr("Sesiones iniciadas", "Calls started"), tr("Incluye históricos dentro del filtro", "Includes historical records within the filter")),
        ("measured", tr("Con telemetría", "With complete tracking"), tr("Sesiones de esta versión", "Calls recorded by this version")),
        ("legacy", tr("Históricos sin telemetría", "Historical records with limited data"), tr("No usados en tasas de respuesta", "Not used for answer rates")),
        ("attempted", tr("INVITE cliente enviados", "Customer calls sent"), tr("Intentos de marcación observados", "Observed call attempts")),
        ("answered", tr("Respuestas cliente", "Customer answers"), tr("SIP 2xx/confirmado, incluye buzones", "Confirmed answers, including voicemail")),
        ("transfer_requested", tr("Solicitudes de agente", "Agent requests"), tr("Opción 2 aceptada", "Option 2 accepted")),
        ("bridged", tr("Puentes con agente", "Agent conversations"), tr("Audio bidireccional establecido", "Two-way audio connected")),
        ("message_completed", tr("Mensajes completos", "Messages completed"), tr("Reproducciones que alcanzaron su final", "Messages played to completion")),
    ]:
        append(main, [label, counts[key], definition])
    append(
        main,
        [
            tr("Tasa de respuesta (%)", "Answer rate (%)"),
            None if summary["answer_rate"] is None else summary["answer_rate"] * 100,
            tr("Respuestas / INVITE cliente observados", "Answers / observed customer calls"),
        ],
    )
    append(
        main,
        [
            tr("Transferencias conectadas (%)", "Connected transfers (%)"),
            None if summary["transfer_rate"] is None else summary["transfer_rate"] * 100,
            tr("Puentes / solicitudes de agente", "Agent conversations / agent requests"),
        ],
    )
    for key, label in [
        ("customer_connected_seconds", tr("Cliente conectado", "Customer connected")),
        ("agent_connected_seconds", tr("Agente conectado", "Agent connected")),
        ("bridge_seconds", tr("Conversación con agente", "Agent conversation")),
    ]:
        metric = summary["durations"].get(key, {})
        append(
            main,
            [
                f"{tr('Promedio', 'Average')} {label.lower()} (s)",
                metric.get("average"),
                f"{metric.get('samples', 0)} {tr('mediciones con fin observado', 'completed measurements')}",
            ],
        )
    append(
        main,
        [
            tr("Nota", "Note"),
            tr("Celdas vacías = no observado", "Blank cells = not observed"),
            tr("Consulta Definiciones antes de interpretar datos.", "Review Definitions before interpreting the data."),
        ],
    )

    trend = sheet(
        tr("Tendencia", "Trend"), [tr("Fecha local", "Local date"), tr("Sesiones", "Calls"), tr("Respuestas", "Answers"), tr("Puentes", "Agent conversations")], [24, 20, 20, 20]
    )
    for day in summary["daily"]:
        append(
            trend,
            [datetime.fromisoformat(day["date"]), day["total"], day["answered"], day["bridged"]],
        )
    if summary["daily"]:
        chart = LineChart()
        chart.title = tr("Actividad de llamadas", "Call activity")
        chart.y_axis.title = tr("Llamadas", "Calls")
        chart.x_axis.title = tr("Fecha local", "Local date")
        chart.style = 13
        chart.width, chart.height = 25, 12
        chart.add_data(
            Reference(trend, min_col=2, max_col=4, min_row=1, max_row=len(summary["daily"]) + 1),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(trend, min_col=1, min_row=2, max_row=len(summary["daily"]) + 1)
        )
        main.add_chart(chart, "E5")
    outcomes = sheet(tr("Resultados", "Outcomes"), [tr("Resultado", "Outcome"), tr("Llamadas", "Calls"), tr("Tipo", "Type")], [32, 20, 32])
    for group, labels, name in [
        ("outcomes", STATUS_LABELS, tr("Resultado operativo", "Call outcome")),
        ("amd", AMD_LABELS, tr("Análisis AMD", "Answer classification")),
        ("hangup_actors", ACTOR_LABELS, tr("Fin de sesión", "Call ended by")),
    ]:
        for key, value in summary[group].items():
            append(outcomes, [_label(labels.get(key, key), language), value, name])
    if summary["outcomes"]:
        chart = BarChart()
        chart.type, chart.style = "bar", 13
        chart.title = tr("Resultados operativos", "Call outcomes")
        chart.width, chart.height = 25, 12
        chart.add_data(
            Reference(outcomes, min_col=2, min_row=1, max_row=len(summary["outcomes"]) + 1),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(outcomes, min_col=1, min_row=2, max_row=len(summary["outcomes"]) + 1)
        )
        main.add_chart(chart, "E29")
    campaigns = sheet(
        tr("Campañas", "Campaigns"), ["ID", tr("Campaña", "Campaign"), tr("Modo", "Operation type"), tr("Sesiones", "Calls"), tr("Respuestas", "Answers"), tr("Puentes", "Agent conversations"), tr("Buzón probable", "Probable voicemail")]
    )
    for item in summary["campaigns"]:
        append(
            campaigns,
            [
                _label(item[key], language) if key == "mode" else item[key]
                for key in ("id", "name", "mode", "total", "answered", "bridged", "machine")
            ],
        )
    cdrs = sheet("CDRs", [CDR_COLUMNS_EN.get(label, label) if english else label for _, label in CDR_COLUMNS])
    for row in rows:
        append(cdrs, [date_value(key, _label(row.get(key), language) if key in {"status_label", "amd_label", "end_actor_label", "mode", "coverage"} else row.get(key)) for key, _ in CDR_COLUMNS])
    from openpyxl.utils import get_column_letter

    cdrs.auto_filter.ref = f"A1:{get_column_letter(len(CDR_COLUMNS))}{len(rows) + 1}"
    legs = sheet(tr("Tramos", "Call legs"), [tr("ID llamada", "Call ID"), tr("Rol", "Role"), tr("Nombre troncal", "Provider name")] + list(LEG_FIELDS))
    for row in rows:
        if "_legs" in row:
            for leg in row["_legs"]:
                append(
                    legs,
                    [row["id"], leg["role"], leg.get("trunk_name")]
                    + [date_value(key, leg.get(key)) for key in LEG_FIELDS],
                )
            continue
        for role in ("customer", "agent"):
            if row[f"{role}_id"]:
                append(
                    legs,
                    [row["id"], role, row.get(f"{role}_trunk_name")]
                    + [date_value(key, row[f"{role}_{key}"]) for key in LEG_FIELDS],
                )
    timeline = sheet(
        tr("Eventos", "Events"), [tr("ID evento", "Event ID"), tr("ID llamada", "Call ID"), tr("ID tramo", "Call leg ID"), tr("Evento", "Event"), tr("Fecha local", "Local date"), tr("Datos JSON", "Event data")]
    )
    for event in events:
        append(
            timeline,
            [
                event["id"],
                event["job_id"],
                event["leg_id"],
                event["kind"],
                date_value("created_at", event["created_at"]),
                event["data"],
            ],
        )
    definitions = sheet(tr("Definiciones", "Definitions"), [tr("Campo / indicador", "Field / metric"), tr("Interpretación", "Interpretation")], [36, 110])
    for definition in DEFINITIONS_EN if english else DEFINITIONS:
        append(definitions, definition)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
